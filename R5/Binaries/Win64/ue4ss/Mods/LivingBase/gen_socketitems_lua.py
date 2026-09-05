# gen_socketitems_lua.py -- (2026-09-07) regenerates the Config.SOCKETITEMS_* block in config.lua
# from RedFalcon's own Other/SocketItems.xlsx (5 tabs: Sockets, Item Ratios, Rarity Ratios, Items,
# Weapons -- see WINDROSE_MODDING_NOTES.md 19u for the full design/rules this data drives).
#
# Run this after ANY edit to SocketItems.xlsx, then paste the printed block over the existing
# Config.SOCKETITEMS_* section in config.lua (it starts right before the "OPTIONAL: ModSettings"
# trailer near the end of the file -- search for "Config.SOCKETITEMS_SOCKETS").
#
# Usage:  python gen_socketitems_lua.py
#
# If Excel has the workbook open, this script cannot read it directly (PermissionError) -- close
# Excel first, or this script will fall back to reading a same-folder copy named
# "SocketItems_copy.xlsx" if one exists.

import openpyxl, re, os, shutil, sys

XLSX_PATH = r"H:\OneDrive\Coding\WINDROSE MODS\Other\SocketItems.xlsx"

def open_workbook(path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except PermissionError:
        fallback = os.path.join(os.path.dirname(path), "SocketItems_copy.xlsx")
        try:
            shutil.copy(path, fallback)
        except Exception:
            print(f"ERROR: '{path}' is locked (probably open in Excel) and no fallback copy "
                  f"could be made. Close Excel and re-run.", file=sys.stderr)
            raise
        print(f"NOTE: source was locked, read a fresh copy at '{fallback}' instead.")
        return openpyxl.load_workbook(fallback, data_only=True)

wb = open_workbook(XLSX_PATH)

def luastr(s):
    if s is None:
        return '""'
    s = str(s).replace("\\", "\\\\").replace('"', '\\"')
    return '"' + s + '"'

def split_list(s):
    if not s:
        return []
    parts = [p.strip() for p in str(s).split(",")]
    return [p for p in parts if p]

def luaarr(items):
    if not items:
        return "{}"
    return "{ " + ", ".join(luastr(i) for i in items) + " }"

out = []

# ---- Sockets ----
ws = wb["Sockets"]
out.append("Config.SOCKETITEMS_SOCKETS = {")
for row in ws.iter_rows(min_row=2, values_only=True):
    socket, locDesc, socType, beltpiece, locTag = row[:5]
    if socket is None:
        continue
    bp = split_list(beltpiece)
    out.append(f"  {{ socket={luastr(socket)}, location={luastr(locDesc)}, socType={luastr(socType)}, beltpiece={luaarr(bp)}, locationTag={luastr(locTag)} }},")
out.append("}")
out.append("")

# ---- Item Ratios ----
ws = wb["Item Ratios"]
out.append("Config.SOCKETITEMS_RATIOS = {")
for row in ws.iter_rows(min_row=2, values_only=True):
    location, type_, count, notes = row[:4]
    if location is None:
        continue
    mandatory = "true" if (notes and "always" in str(notes).lower()) else "false"
    out.append(f"  {{ location={luastr(location)}, type={luastr(type_)}, count={int(count)}, mandatory={mandatory} }},{'  -- ' + str(notes) if notes else ''}")
out.append("}")
out.append("")

# ---- Rarity Ratios ----
ws = wb["Rarity Ratios"]
out.append("Config.SOCKETITEMS_RARITY_WEIGHTS = {")
for row in ws.iter_rows(min_row=2, values_only=True):
    rarity, chance = row[:2]
    if rarity is None:
        continue
    out.append(f"  [{luastr(rarity)}] = {chance},")
out.append("}")
out.append("")

# ---- Items ----
ws = wb["Items"]
out.append("Config.SOCKETITEMS_ITEMS = {")
for row in ws.iter_rows(min_row=2, values_only=True):
    asset, shortName, availSocket, limit, tag, rarity = row[:6]
    if asset is None:
        continue
    socks = split_list(availSocket)
    tags = split_list(tag)
    # Defensive Limit parse -- one real row has had a stray string like "1_L" instead of a number.
    if isinstance(limit, (int, float)):
        limitNum = int(limit)
    else:
        m = re.match(r"^\s*(\d+)", str(limit or ""))
        limitNum = int(m.group(1)) if m else 1
    out.append(f"  {{ asset={luastr(asset)}, shortName={luastr(shortName)}, sockets={luaarr(socks)}, limit={limitNum}, tags={luaarr(tags)}, rarity={luastr(rarity)} }},")
out.append("}")
out.append("")

# ---- Weapons ----
ws = wb["Weapons"]
out.append("Config.SOCKETITEMS_WEAPONS = {")
for row in ws.iter_rows(min_row=2, values_only=True):
    asset, shortName, availSocket, tag, location, rarity = row[:6]
    if asset is None:
        continue
    socks = split_list(availSocket)
    tags = split_list(tag)
    out.append(f"  {{ asset={luastr(asset)}, shortName={luastr(shortName)}, sockets={luaarr(socks)}, tags={luaarr(tags)}, location={luastr(location)}, rarity={luastr(rarity)} }},")
out.append("}")

result = "\n".join(out)
OUT_PATH = os.path.join(os.path.dirname(__file__), "socketitems_generated.lua")
with open(OUT_PATH, "w", encoding="utf-8") as f:
    f.write(result)
print(f"wrote {OUT_PATH}, {len(result)} chars -- paste this over the existing Config.SOCKETITEMS_* block in config.lua")
